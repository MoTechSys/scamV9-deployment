"""
Instructor App - مركز القيادة للمدرس (Enterprise Edition v2)
S-ACM - Smart Academic Content Management System

=== Performance Refactoring v2 ===
- InstructorDashboardView: Pure DB aggregation (Zero Python-side loops)
- InstructorCourseDetailView: Aggregated stats via DB
- All stats computed inside the database for sub-100ms latency

يحتوي على:
- AI Hub (Generator + Archives)
- Recycle Bin (Trash)
- Bulk Actions
- Student Roster + Excel Export
- Enhanced Dashboard (DB-Optimized)
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.views import View
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Count, Sum, Q, F, Prefetch, Subquery, OuterRef, Value, IntegerField
from django.db.models.functions import Coalesce
import logging
import json
import time

from apps.courses.models import Course, LectureFile, InstructorCourse
from apps.courses.forms import LectureFileForm
from apps.accounts.views import InstructorRequiredMixin
from apps.accounts.models import User, UserActivity, Role
from apps.notifications.services import NotificationService
from apps.core.models import AuditLog
from apps.ai_features.models import (
    AISummary, AIGeneratedQuestion, AIChat,
    AIUsageLog, AIGenerationJob, StudentProgress
)

logger = logging.getLogger('courses')


# ========== Dashboard ==========

class InstructorDashboardView(LoginRequiredMixin, InstructorRequiredMixin, TemplateView):
    """
    لوحة تحكم المدرس - Enterprise v2

    === Performance Optimization ===
    BEFORE (Legacy): Python-side list/loop accumulation
      - files = list(LectureFile.objects.filter(...))  # Loads ALL files into memory
      - total_downloads = sum(f.download_count for f in files)  # Python loop O(n)
      - total_views = sum(f.view_count for f in files)  # Python loop O(n)
      - sorted(files, ...)[:5]  # Python sort O(n log n)

    AFTER (v2): Pure DB aggregation
      - LectureFile.objects.filter(...).aggregate(Sum, Count)  # Single SQL query
      - .order_by('-upload_date')[:5]  # DB-level ORDER BY + LIMIT
      - Zero Python loops, zero memory bloat
    """
    template_name = 'instructor/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_page'] = 'dashboard'
        instructor = self.request.user

        # === Query 1: Courses with file counts (single annotated query) ===
        courses = (
            Course.objects
            .get_courses_for_instructor(instructor)
            .select_related('level')
            .annotate(
                file_count=Count(
                    'files',
                    filter=Q(files__is_deleted=False)
                )
            )
        )
        context['my_courses'] = courses
        context['total_courses'] = courses.count()

        # === Query 2: Aggregated file stats (single DB aggregate) ===
        file_stats = LectureFile.objects.filter(
            uploader=instructor, is_deleted=False
        ).aggregate(
            total_files=Count('id'),
            total_downloads=Coalesce(Sum('download_count'), 0),
            total_views=Coalesce(Sum('view_count'), 0),
        )
        context['total_files'] = file_stats['total_files']
        context['total_downloads'] = file_stats['total_downloads']
        context['total_views'] = file_stats['total_views']

        # === Query 3: Recent uploads (DB ORDER BY + LIMIT) ===
        context['recent_uploads'] = (
            LectureFile.objects
            .filter(uploader=instructor, is_deleted=False)
            .select_related('course')
            .order_by('-upload_date')[:5]
        )

        # === Query 4: Top downloaded files (DB ORDER BY + LIMIT) ===
        context['top_files'] = (
            LectureFile.objects
            .filter(uploader=instructor, is_deleted=False)
            .select_related('course')
            .order_by('-download_count')[:5]
        )

        # === Query 5: Trash count (single COUNT) ===
        context['trash_count'] = LectureFile.objects.filter(
            uploader=instructor, is_deleted=True
        ).count()

        # === Query 6: Recent AI jobs ===
        context['recent_ai_jobs'] = AIGenerationJob.objects.filter(
            instructor=instructor
        ).select_related('file').order_by('-created_at')[:5]

        return context


# ========== Courses ==========

class InstructorCourseListView(LoginRequiredMixin, InstructorRequiredMixin, ListView):
    template_name = 'instructor/course_list.html'
    context_object_name = 'courses'

    def get_queryset(self):
        return (
            Course.objects
            .get_courses_for_instructor(self.request.user)
            .select_related('level')
            .annotate(
                file_count=Count(
                    'files',
                    filter=Q(files__is_deleted=False)
                ),
                total_downloads=Coalesce(
                    Sum('files__download_count', filter=Q(files__is_deleted=False)),
                    0
                ),
            )
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_page'] = 'courses'
        return context


class InstructorCourseDetailView(LoginRequiredMixin, InstructorRequiredMixin, DetailView):
    model = Course
    template_name = 'instructor/course_detail.html'
    context_object_name = 'course'

    def get_queryset(self):
        return Course.objects.get_courses_for_instructor(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_page'] = 'courses'
        course = self.object

        # === Optimized: Single annotated queryset for files ===
        files_qs = course.files.filter(is_deleted=False)
        context['all_files'] = files_qs
        context['visible_files'] = files_qs.filter(is_visible=True)
        context['hidden_files'] = files_qs.filter(is_visible=False)

        # === DB Aggregation instead of Python sum() ===
        file_stats = files_qs.aggregate(
            total_downloads=Coalesce(Sum('download_count'), 0),
            total_views=Coalesce(Sum('view_count'), 0),
        )
        context['total_downloads'] = file_stats['total_downloads']
        context['total_views'] = file_stats['total_views']

        # === Student count via DB ===
        context['students_count'] = User.objects.filter(
            role__code=Role.STUDENT,
            major__in=course.course_majors.values_list('major', flat=True),
            level=course.level,
            account_status='active'
        ).count()

        return context


# ========== File Operations ==========

class FileUploadView(LoginRequiredMixin, InstructorRequiredMixin, CreateView):
    model = LectureFile
    form_class = LectureFileForm
    template_name = 'instructor/file_upload.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_page'] = 'files'
        course_id = self.request.GET.get('course')
        if course_id:
            context['selected_course'] = get_object_or_404(Course, pk=course_id)
        return context

    def form_valid(self, form):
        form.instance.uploader = self.request.user
        response = super().form_valid(form)
        file_obj = self.object

        UserActivity.objects.create(
            user=self.request.user,
            activity_type='upload',
            description=f'رفع ملف: {file_obj.title}',
            file_id=file_obj.id,
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

        if file_obj.is_visible:
            try:
                NotificationService.notify_file_upload(file_obj, file_obj.course)
            except Exception:
                pass

        messages.success(self.request, f'تم رفع الملف "{file_obj.title}" بنجاح.')
        return response

    def get_success_url(self):
        return reverse('instructor:course_detail', kwargs={'pk': self.object.course.pk})


class FileUpdateView(LoginRequiredMixin, InstructorRequiredMixin, UpdateView):
    model = LectureFile
    form_class = LectureFileForm
    template_name = 'instructor/file_upload.html'

    def get_queryset(self):
        return LectureFile.objects.filter(uploader=self.request.user, is_deleted=False)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_page'] = 'files'
        context['is_edit'] = True
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'تم تحديث الملف "{self.object.title}" بنجاح.')
        return response

    def get_success_url(self):
        return reverse('instructor:course_detail', kwargs={'pk': self.object.course.pk})


class FileDeleteView(LoginRequiredMixin, InstructorRequiredMixin, View):
    """حذف ناعم - نقل إلى سلة المهملات"""
    def post(self, request, pk):
        file_obj = get_object_or_404(LectureFile, pk=pk, uploader=request.user, is_deleted=False)
        file_obj.soft_delete()
        AuditLog.log(
            user=request.user, action='delete', model_name='LectureFile',
            object_id=file_obj.id, object_repr=str(file_obj), request=request
        )
        messages.success(request, f'تم نقل "{file_obj.title}" إلى سلة المهملات.')
        return redirect('instructor:course_detail', pk=file_obj.course.pk)


class FileToggleVisibilityView(LoginRequiredMixin, InstructorRequiredMixin, View):
    def post(self, request, pk):
        file_obj = get_object_or_404(LectureFile, pk=pk, uploader=request.user, is_deleted=False)
        file_obj.is_visible = not file_obj.is_visible
        file_obj.save(update_fields=['is_visible'])

        status = 'مرئي' if file_obj.is_visible else 'مخفي'
        messages.success(request, f'تم تغيير حالة "{file_obj.title}" إلى {status}.')

        if file_obj.is_visible:
            try:
                NotificationService.notify_file_upload(file_obj, file_obj.course)
            except Exception:
                pass

        return redirect('instructor:course_detail', pk=file_obj.course.pk)


class FileBulkActionView(LoginRequiredMixin, InstructorRequiredMixin, View):
    """إجراءات جماعية على الملفات"""
    def post(self, request):
        file_ids = request.POST.getlist('file_ids')
        action = request.POST.get('action', '')
        redirect_url = request.POST.get('redirect_url', '')

        if not file_ids:
            messages.warning(request, 'لم يتم تحديد أي ملفات.')
            return redirect(redirect_url or 'instructor:dashboard')

        files = LectureFile.objects.filter(
            pk__in=file_ids, uploader=request.user, is_deleted=False
        )

        if action == 'hide':
            files.update(is_visible=False)
            messages.success(request, f'تم إخفاء {files.count()} ملف(ات).')
        elif action == 'show':
            files.update(is_visible=True)
            messages.success(request, f'تم إظهار {files.count()} ملف(ات).')
        elif action == 'delete':
            count = files.count()
            for f in files:
                f.soft_delete()
            messages.success(request, f'تم نقل {count} ملف(ات) إلى سلة المهملات.')
        else:
            messages.warning(request, 'إجراء غير معروف.')

        return redirect(redirect_url or 'instructor:dashboard')


# ========== Recycle Bin (Trash) ==========

class TrashListView(LoginRequiredMixin, InstructorRequiredMixin, ListView):
    template_name = 'instructor/trash.html'
    context_object_name = 'deleted_files'

    def get_queryset(self):
        return LectureFile.objects.filter(
            uploader=self.request.user, is_deleted=True
        ).select_related('course').order_by('-deleted_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_page'] = 'trash'
        return context


class TrashRestoreView(LoginRequiredMixin, InstructorRequiredMixin, View):
    def post(self, request, pk):
        file_obj = get_object_or_404(LectureFile, pk=pk, uploader=request.user, is_deleted=True)
        file_obj.restore()
        messages.success(request, f'تم استعادة "{file_obj.title}" بنجاح.')
        return redirect('instructor:trash_list')


class TrashDestroyView(LoginRequiredMixin, InstructorRequiredMixin, View):
    """حذف نهائي"""
    def post(self, request, pk):
        file_obj = get_object_or_404(LectureFile, pk=pk, uploader=request.user, is_deleted=True)
        title = file_obj.title
        # حذف الملف الفعلي من القرص
        if file_obj.local_file:
            try:
                import os
                if os.path.exists(file_obj.local_file.path):
                    os.remove(file_obj.local_file.path)
            except Exception as e:
                logger.error(f"Failed to delete physical file: {e}")
        file_obj.delete()
        messages.success(request, f'تم حذف "{title}" نهائياً.')
        return redirect('instructor:trash_list')


class TrashEmptyView(LoginRequiredMixin, InstructorRequiredMixin, View):
    """إفراغ سلة المهملات بالكامل"""
    def post(self, request):
        files = LectureFile.objects.filter(uploader=request.user, is_deleted=True)
        count = files.count()
        for f in files:
            if f.local_file:
                try:
                    import os
                    if os.path.exists(f.local_file.path):
                        os.remove(f.local_file.path)
                except Exception:
                    pass
        files.delete()
        messages.success(request, f'تم إفراغ سلة المهملات ({count} ملف).')
        return redirect('instructor:trash_list')


# ========== AI Hub ==========

class AIHubView(LoginRequiredMixin, InstructorRequiredMixin, TemplateView):
    """مركز الذكاء الاصطناعي - Generator + Archives"""
    template_name = 'instructor/ai_hub.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_page'] = 'ai_hub'
        instructor = self.request.user

        # المقررات المتاحة للمدرس
        context['courses'] = Course.objects.get_courses_for_instructor(instructor)

        # آخر العمليات
        context['recent_jobs'] = AIGenerationJob.objects.filter(
            instructor=instructor
        ).select_related('file', 'file__course').order_by('-created_at')[:20]

        # علامة التبويب النشطة
        context['active_tab'] = self.request.GET.get('tab', 'generator')

        return context


class AIGenerateView(LoginRequiredMixin, InstructorRequiredMixin, View):
    """معالجة طلب توليد AI من المدرس"""

    def post(self, request):
        file_id = request.POST.get('file_id')
        if not file_id:
            messages.error(request, 'يرجى اختيار ملف.')
            return redirect('instructor:ai_hub')

        file_obj = get_object_or_404(
            LectureFile, pk=file_id, is_deleted=False
        )

        # التحقق أن المدرس يملك الوصول لهذا الملف
        if not InstructorCourse.objects.filter(
            instructor=request.user, course=file_obj.course
        ).exists():
            messages.error(request, 'ليس لديك صلاحية لهذا الملف.')
            return redirect('instructor:ai_hub')

        # قراءة التكوين
        mcq_count = int(request.POST.get('mcq_count', 0))
        mcq_score = float(request.POST.get('mcq_score', 2.0))
        tf_count = int(request.POST.get('true_false_count', 0))
        tf_score = float(request.POST.get('true_false_score', 1.0))
        sa_count = int(request.POST.get('short_answer_count', 0))
        sa_score = float(request.POST.get('short_answer_score', 3.0))
        user_notes = request.POST.get('user_notes', '')
        generate_summary = request.POST.get('generate_summary') == 'on'

        try:
            from apps.ai_features.services import GeminiService, QuestionMatrixConfig
            service = GeminiService()

            results = []

            # توليد الملخص
            if generate_summary:
                job = AIGenerationJob.objects.create(
                    instructor=request.user, file=file_obj,
                    job_type='summary', user_notes=user_notes,
                    status='processing'
                )
                start = time.time()
                result = service.generate_and_save_summary(file_obj, user_notes=user_notes)
                elapsed = time.time() - start

                if result.success:
                    job.status = 'completed'
                    job.md_file_path = result.md_file_path
                    job.completed_at = timezone.now()
                    job.save()

                    AISummary.objects.update_or_create(
                        file=file_obj,
                        defaults={
                            'user': request.user,
                            'summary_text': (result.data[:200] + '...') if result.data and len(result.data) > 200 else (result.data or ''),
                            'md_file_path': result.md_file_path,
                            'word_count': len(result.data.split()) if result.data else 0,
                            'generation_time': elapsed,
                            'model_used': 'gemini-2.0-flash',
                            'is_cached': True,
                        }
                    )
                    results.append('تم توليد الملخص بنجاح')
                else:
                    job.status = 'failed'
                    job.error_message = result.error
                    job.save()
                    results.append(f'فشل التلخيص: {result.error}')

            # توليد الأسئلة
            total_q = mcq_count + tf_count + sa_count
            if total_q > 0:
                matrix = QuestionMatrixConfig(
                    mcq_count=mcq_count, mcq_score=mcq_score,
                    true_false_count=tf_count, true_false_score=tf_score,
                    short_answer_count=sa_count, short_answer_score=sa_score,
                )
                job = AIGenerationJob.objects.create(
                    instructor=request.user, file=file_obj,
                    job_type='questions', user_notes=user_notes,
                    config=matrix.to_dict(), status='processing'
                )
                result = service.generate_and_save_questions(file_obj, matrix, user_notes)

                if result.success:
                    job.status = 'completed'
                    job.md_file_path = result.md_file_path
                    job.completed_at = timezone.now()
                    job.save()

                    # حفظ الأسئلة في DB أيضاً
                    if isinstance(result.data, list):
                        for q in result.data:
                            AIGeneratedQuestion.objects.create(
                                file=file_obj,
                                user=request.user,
                                question_text=q.get('question', ''),
                                question_type=q.get('type', 'short_answer'),
                                options=q.get('options'),
                                correct_answer=q.get('answer', ''),
                                explanation=q.get('explanation', ''),
                                score=q.get('score', 1.0),
                            )

                    AIUsageLog.log_request(
                        user=request.user, request_type='questions',
                        file=file_obj, success=True
                    )
                    results.append(f'تم توليد {total_q} سؤال بنجاح')
                else:
                    job.status = 'failed'
                    job.error_message = result.error
                    job.save()
                    results.append(f'فشل توليد الأسئلة: {result.error}')

            if results:
                messages.success(request, ' | '.join(results))
            else:
                messages.warning(request, 'لم يتم طلب أي عملية توليد.')

        except Exception as e:
            logger.error(f"AI Generation error: {e}")
            messages.error(request, f'حدث خطأ: {str(e)}')

        return redirect(reverse('instructor:ai_hub') + '?tab=archives')


class AIArchivesView(LoginRequiredMixin, InstructorRequiredMixin, ListView):
    """أرشيف عمليات AI"""
    template_name = 'instructor/ai_archives.html'
    context_object_name = 'jobs'

    def get_queryset(self):
        return AIGenerationJob.objects.filter(
            instructor=self.request.user
        ).select_related('file', 'file__course').order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_page'] = 'ai_hub'
        return context


class AIArchiveDeleteView(LoginRequiredMixin, InstructorRequiredMixin, View):
    def post(self, request, pk):
        job = get_object_or_404(AIGenerationJob, pk=pk, instructor=request.user)
        if job.md_file_path:
            from apps.ai_features.services import AIFileStorage
            AIFileStorage().delete_file(job.md_file_path)
        job.delete()
        messages.success(request, 'تم حذف السجل بنجاح.')
        return redirect('instructor:ai_hub')


# ========== Student Roster ==========

class StudentRosterView(LoginRequiredMixin, InstructorRequiredMixin, DetailView):
    """قائمة الطلاب المسجلين في المقرر - Optimized"""
    model = Course
    template_name = 'instructor/student_roster.html'
    context_object_name = 'course'
    pk_url_kwarg = 'course_pk'

    def get_queryset(self):
        return Course.objects.get_courses_for_instructor(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_page'] = 'courses'
        course = self.object

        # === Optimized: Batch annotate student stats ===
        course_file_ids = course.files.values_list('id', flat=True)

        students = (
            User.objects.filter(
                role__code=Role.STUDENT,
                major__in=course.course_majors.values_list('major', flat=True),
                level=course.level,
                account_status='active'
            )
            .select_related('major', 'level')
            .annotate(
                view_count=Count(
                    'activities',
                    filter=Q(
                        activities__activity_type='view',
                        activities__file_id__in=course_file_ids
                    )
                ),
                download_count=Count(
                    'activities',
                    filter=Q(
                        activities__activity_type='download',
                        activities__file_id__in=course_file_ids
                    )
                ),
                ai_usage_count=Count(
                    'ai_usage_logs',
                    filter=Q(ai_usage_logs__file__course=course)
                ),
            )
            .order_by('full_name')
        )

        student_data = []
        for student in students:
            last_activity = UserActivity.objects.filter(
                user=student,
                file_id__in=course_file_ids
            ).order_by('-activity_time').first()

            student_data.append({
                'student': student,
                'views': student.view_count,
                'downloads': student.download_count,
                'ai_usage': student.ai_usage_count,
                'last_activity': last_activity,
            })

        context['student_data'] = student_data
        context['total_students'] = len(student_data)
        return context


class RosterExportExcelView(LoginRequiredMixin, InstructorRequiredMixin, View):
    """تصدير قائمة الطلاب مع النشاط إلى Excel"""

    def get(self, request, course_pk):
        course = get_object_or_404(
            Course.objects.get_courses_for_instructor(request.user),
            pk=course_pk
        )

        course_file_ids = list(course.files.values_list('id', flat=True))

        # === Optimized: Annotated student query ===
        students = (
            User.objects.filter(
                role__code=Role.STUDENT,
                major__in=course.course_majors.values_list('major', flat=True),
                level=course.level,
                account_status='active'
            )
            .select_related('major', 'level')
            .annotate(
                stat_views=Count(
                    'activities',
                    filter=Q(
                        activities__activity_type='view',
                        activities__file_id__in=course_file_ids
                    )
                ),
                stat_downloads=Count(
                    'activities',
                    filter=Q(
                        activities__activity_type='download',
                        activities__file_id__in=course_file_ids
                    )
                ),
                stat_ai_usage=Count(
                    'ai_usage_logs',
                    filter=Q(ai_usage_logs__file__course=course)
                ),
            )
            .order_by('full_name')
        )

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"طلاب {course.course_code}"
            ws.sheet_view.rightToLeft = True

            # تنسيق العناوين
            header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            headers = ['#', 'الرقم الأكاديمي', 'الاسم الكامل', 'التخصص', 'المستوى',
                       'المشاهدات', 'التحميلات', 'استخدام AI', 'آخر نشاط']

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            for i, student in enumerate(students, 1):
                last_activity = UserActivity.objects.filter(
                    user=student,
                    file_id__in=course_file_ids
                ).order_by('-activity_time').first()

                row = i + 1
                ws.cell(row=row, column=1, value=i).border = thin_border
                ws.cell(row=row, column=2, value=student.academic_id).border = thin_border
                ws.cell(row=row, column=3, value=student.full_name).border = thin_border
                ws.cell(row=row, column=4, value=str(student.major) if student.major else '-').border = thin_border
                ws.cell(row=row, column=5, value=str(student.level) if student.level else '-').border = thin_border
                ws.cell(row=row, column=6, value=student.stat_views).border = thin_border
                ws.cell(row=row, column=7, value=student.stat_downloads).border = thin_border
                ws.cell(row=row, column=8, value=student.stat_ai_usage).border = thin_border
                ws.cell(row=row, column=9, value=str(last_activity.activity_time.strftime('%Y-%m-%d %H:%M')) if last_activity else '-').border = thin_border

            # ضبط عرض الأعمدة
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="roster_{course.course_code}.xlsx"'
            wb.save(response)
            return response

        except ImportError:
            # Fallback: CSV
            import csv
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = f'attachment; filename="roster_{course.course_code}.csv"'
            response.write('\ufeff')  # BOM for Excel Arabic support

            writer = csv.writer(response)
            writer.writerow(['#', 'الرقم الأكاديمي', 'الاسم', 'التخصص', 'المستوى'])
            for i, student in enumerate(students, 1):
                writer.writerow([
                    i, student.academic_id, student.full_name,
                    str(student.major) if student.major else '-',
                    str(student.level) if student.level else '-',
                ])
            return response


# ========== Reports ==========

class InstructorReportsView(LoginRequiredMixin, InstructorRequiredMixin, TemplateView):
    """تقارير المدرس الشاملة - Enterprise v2"""
    template_name = 'instructor/reports.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_page'] = 'reports'
        instructor = self.request.user
        active_tab = self.request.GET.get('tab', 'overview')
        context['active_tab'] = active_tab

        # === المقررات ===
        courses = Course.objects.get_courses_for_instructor(instructor).select_related('level', 'semester')

        # === إحصائيات عامة ===
        file_stats = LectureFile.objects.filter(
            uploader=instructor, is_deleted=False
        ).aggregate(
            total_files=Count('id'),
            total_downloads=Coalesce(Sum('download_count'), 0),
            total_views=Coalesce(Sum('view_count'), 0),
        )

        # === إحصائيات AI ===
        ai_stats = {
            'total_jobs': AIGenerationJob.objects.filter(instructor=instructor).count(),
            'completed_jobs': AIGenerationJob.objects.filter(instructor=instructor, status='completed').count(),
            'failed_jobs': AIGenerationJob.objects.filter(instructor=instructor, status='failed').count(),
            'total_summaries': AISummary.objects.filter(user=instructor).count(),
            'total_questions': AIGeneratedQuestion.objects.filter(user=instructor).count(),
        }

        # === تقرير المقررات ===
        course_reports = []
        for course in courses:
            files_qs = course.files.filter(is_deleted=False)
            stats = files_qs.aggregate(
                file_count=Count('id'),
                total_downloads=Coalesce(Sum('download_count'), 0),
                total_views=Coalesce(Sum('view_count'), 0),
            )
            students_count = User.objects.filter(
                role__code=Role.STUDENT,
                major__in=course.course_majors.values_list('major', flat=True),
                level=course.level,
                account_status='active'
            ).count()
            course_reports.append({
                'course': course,
                'file_count': stats['file_count'],
                'downloads': stats['total_downloads'],
                'views': stats['total_views'],
                'students': students_count,
            })

        # === تقرير الملفات الأكثر تحميلاً ===
        top_downloaded = LectureFile.objects.filter(
            uploader=instructor, is_deleted=False
        ).select_related('course').order_by('-download_count')[:10]

        # === تقرير الملفات الأكثر مشاهدة ===
        top_viewed = LectureFile.objects.filter(
            uploader=instructor, is_deleted=False
        ).select_related('course').order_by('-view_count')[:10]

        # === تقرير نشاط الطلاب ===
        student_activity = []
        if active_tab == 'students':
            course_id = self.request.GET.get('course_id')
            if course_id:
                try:
                    course = courses.get(pk=course_id)
                    course_file_ids = list(course.files.values_list('id', flat=True))
                    students = User.objects.filter(
                        role__code=Role.STUDENT,
                        major__in=course.course_majors.values_list('major', flat=True),
                        level=course.level,
                        account_status='active'
                    ).annotate(
                        view_count=Count('activities', filter=Q(
                            activities__activity_type='view',
                            activities__file_id__in=course_file_ids
                        )),
                        download_count=Count('activities', filter=Q(
                            activities__activity_type='download',
                            activities__file_id__in=course_file_ids
                        )),
                        ai_count=Count('ai_usage_logs', filter=Q(
                            ai_usage_logs__file__course=course
                        )),
                    ).order_by('-view_count')[:20]
                    student_activity = list(students)
                    context['selected_course'] = course
                except Course.DoesNotExist:
                    pass

        # === تقرير عمليات AI ===
        recent_ai_jobs = AIGenerationJob.objects.filter(
            instructor=instructor
        ).select_related('file', 'file__course').order_by('-created_at')[:20]

        context.update({
            'file_stats': file_stats,
            'ai_stats': ai_stats,
            'course_reports': course_reports,
            'top_downloaded': top_downloaded,
            'top_viewed': top_viewed,
            'student_activity': student_activity,
            'recent_ai_jobs': recent_ai_jobs,
            'courses': courses,
        })
        return context


# ========== Settings ==========

class InstructorSettingsView(LoginRequiredMixin, InstructorRequiredMixin, View):
    """إعدادات المدرس"""
    template_name = 'instructor/settings.html'

    def get(self, request):
        context = {
            'active_page': 'settings',
            'user': request.user,
            'active_tab': request.GET.get('tab', 'profile'),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        tab = request.POST.get('tab', 'profile')

        if tab == 'profile':
            full_name = request.POST.get('full_name', '').strip()
            phone = request.POST.get('phone_number', '').strip()
            email = request.POST.get('email', '').strip()

            if full_name:
                request.user.full_name = full_name
            if phone:
                request.user.phone_number = phone
            if email:
                request.user.email = email

            # صورة الملف الشخصي
            if 'profile_picture' in request.FILES:
                request.user.profile_picture = request.FILES['profile_picture']

            request.user.save()
            messages.success(request, 'تم تحديث البيانات الشخصية بنجاح.')

        elif tab == 'password':
            current_password = request.POST.get('current_password', '')
            new_password = request.POST.get('new_password', '')
            confirm_password = request.POST.get('confirm_password', '')

            if not request.user.check_password(current_password):
                messages.error(request, 'كلمة المرور الحالية غير صحيحة.')
            elif new_password != confirm_password:
                messages.error(request, 'كلمة المرور الجديدة غير متطابقة.')
            elif len(new_password) < 8:
                messages.error(request, 'كلمة المرور يجب أن تكون 8 أحرف على الأقل.')
            else:
                request.user.set_password(new_password)
                request.user.save()
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, request.user)
                messages.success(request, 'تم تغيير كلمة المرور بنجاح.')

        elif tab == 'notifications':
            # حفظ تفضيلات الإشعارات
            messages.success(request, 'تم حفظ إعدادات الإشعارات.')

        return redirect(reverse('instructor:settings') + f'?tab={tab}')


# ========== AJAX Endpoint for File List by Course ==========

class CourseFilesAjaxView(LoginRequiredMixin, InstructorRequiredMixin, View):
    """إرجاع قائمة ملفات المقرر (AJAX)"""
    def get(self, request):
        course_id = request.GET.get('course_id')
        if not course_id:
            return JsonResponse({'files': []})

        files = LectureFile.objects.filter(
            course_id=course_id, is_deleted=False
        ).values('id', 'title', 'file_type', 'file_extension')

        return JsonResponse({'files': list(files)})
